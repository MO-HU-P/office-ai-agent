"""Excel (.xlsx) 操作ツール群"""
from copy import copy
from typing import Any, Optional

from langchain_core.tools import tool
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries

from ...atomic import atomic_save
from ...config import resolve_workspace_path
from .inline_format import describe_format, validate_format_args


# グラフの系列色。プレビュー(ExcelChart.tsx)の既定色と同じ並び。
# 隣り合う色が色覚特性のある人にも見分けられるよう検証済みなので、順番は変えないこと
CHART_COLORS = ["2A78D6", "008300", "E87BA4", "EDA100", "1BAF7A", "EB6834", "4A3AA7", "E34948"]


def _absolute(a1_range: str) -> str:
    """"C2:C5" → "$C$2:$C$5"。グラフの参照式は絶対参照で書く必要がある。"""
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    return (f"${get_column_letter(min_col)}${min_row}"
            f":${get_column_letter(max_col)}${max_row}")


def _open(filename: str):
    path = resolve_workspace_path(filename, must_exist=True)
    if path.suffix != ".xlsx":
        raise ValueError("Excelブックは .xlsx ファイルを指定してください")
    return load_workbook(str(path)), str(path)


def _sheet(wb, sheet: Optional[str]):
    if sheet is None:
        return wb.active
    if sheet in wb.sheetnames:
        return wb[sheet]
    return wb.create_sheet(sheet)


def _set_value(ws, ref: str, value: Any):
    cell = ws[ref]
    if isinstance(value, str) and value.startswith("="):
        cell.value = value  # 数式
    else:
        cell.value = value


def _overwrite_warning(refs: list[str]) -> str:
    """既に値があったセルを上書きしたことをモデルに知らせる注意書き。
    モデルが書き込み位置を誤って既存データを消したとき(表の途中に結果を書く等)、
    次のターンで気づいて restore_file でやり直せるようにする。"""
    if not refs:
        return ""
    shown = ", ".join(refs[:8]) + (" ほか" if len(refs) > 8 else "")
    return (f"\n⚠️ 注意: {len(refs)}個のセルには既に値があり、上書きしました({shown})。"
            "意図した更新ならそのまま進めてよい。既存の表を誤って上書きした場合は、"
            "restore_file でこの書き込みの前に戻し、既存データの最終行より下か新しいシートに書き直すこと。")


@tool
def excel_create(filename: str, sheet_name: str = "Sheet1") -> str:
    """新しいExcelブック(.xlsx)を作成する。filenameは必ず.xlsxで終わること。"""
    path = resolve_workspace_path(filename)
    if path.suffix != ".xlsx":
        return "エラー: filenameは .xlsx で終わる必要があります"
    wb = Workbook()
    wb.active.title = sheet_name
    atomic_save(wb.save, path)
    return f"{filename} を作成しました (シート: {sheet_name})"


@tool
def excel_read(filename: str, sheet: str = "", cell_range: str = "", mode: str = "full") -> str:
    """Excelブックの内容を読む。sheet省略時は全シート名一覧+アクティブシートの内容を返す。
    cell_rangeで範囲指定可能(例: "A1:D10")。数式セルは数式のまま表示される。
    mode="summary"にすると各シートの名前と使用範囲だけを返す。大きなブックはまずsummaryで
    全体を把握し、必要なシート・範囲だけを読むこと。"""
    wb, _ = _open(filename)
    if mode == "summary":
        lines = [f"{ws.title}: {ws.max_row}行 x {ws.max_column}列" for ws in wb.worksheets]
        return "シート一覧:\n" + "\n".join(lines)
    lines = [f"シート一覧: {', '.join(wb.sheetnames)}"]
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    lines.append(f"--- シート「{ws.title}」({ws.max_row}行 x {ws.max_column}列) ---")
    if cell_range:
        min_c, min_r, max_c, max_r = range_boundaries(cell_range)
        # 巨大な範囲指定で応答が肥大化しないよう、実データがある範囲+上限に丸める
        max_r = min(max_r, ws.max_row, min_r + 499)
        max_c = min(max_c, ws.max_column, min_c + 59)
    else:
        min_r, min_c = 1, 1
        max_r, max_c = min(ws.max_row, 100), min(ws.max_column, 30)
    for r in range(min_r, max_r + 1):
        cells = []
        for c in range(min_c, max_c + 1):
            v = ws.cell(row=r, column=c).value
            cells.append("" if v is None else str(v))
        if any(cells):
            lines.append(f"{r}: " + " | ".join(cells))
    return "\n".join(lines)


@tool
def excel_write_cells(filename: str, cells: dict[str, Any], sheet: str = "") -> str:
    """Excelのセルに値を書き込む。cellsはセル番地→値の辞書。例: {"A1": "商品名", "B1": 100, "C1": "=A1*B1"}
    "=" で始まる文字列は数式として書き込まれる。sheetが存在しない場合は新規作成される。"""
    wb, path = _open(filename)
    ws = _sheet(wb, sheet or None)
    overwritten = []
    for ref, value in cells.items():
        if ws[ref].value is not None:
            overwritten.append(ref)
        _set_value(ws, ref, value)
    atomic_save(wb.save, path)
    return (f"{filename} のシート「{ws.title}」に {len(cells)} セルを書き込みました"
            + _overwrite_warning(overwritten))


@tool
def excel_write_rows(filename: str, start_cell: str, rows: list[list[Any]], sheet: str = "") -> str:
    """Excelに複数行のデータを一括で書き込む。start_cell(例: "A2")を左上として、rows(2次元配列)を展開する。
    表データを書くときはexcel_write_cellsよりこちらを使うこと。"=" で始まる値は数式。"""
    wb, path = _open(filename)
    ws = _sheet(wb, sheet or None)
    min_c, min_r, _, _ = range_boundaries(start_cell)
    n = 0
    overwritten = []
    for ri, row in enumerate(rows):
        for ci, value in enumerate(row):
            ref = f"{get_column_letter(min_c + ci)}{min_r + ri}"
            if ws[ref].value is not None:
                overwritten.append(ref)
            _set_value(ws, ref, value)
            n += 1
    atomic_save(wb.save, path)
    return (f"{filename} のシート「{ws.title}」に {len(rows)}行 ({n}セル) を書き込みました"
            + _overwrite_warning(overwritten))


@tool
def excel_format(
    filename: str,
    cell_range: str,
    sheet: str = "",
    bold: Optional[bool] = None,
    italic: Optional[bool] = None,
    font_size: Optional[float] = None,
    font_color: str = "",
    bg_color: str = "",
    align: str = "",
    number_format: str = "",
    col_width: Optional[float] = None,
) -> str:
    """Excelのセル範囲に書式を設定する。cell_rangeは "A1:D1" のような範囲か単一セル。
    "A1:D1,A5:D5" のようにカンマ区切りで複数範囲へ同じ書式を一括設定できる(1範囲ずつ繰り返さないこと)。
    font_color/bg_colorは "#RRGGBB" 形式。alignは left/center/right。
    number_formatは "#,##0" "0.0%" "yyyy/mm/dd" など。col_widthで列幅も設定できる。"""
    wb, path = _open(filename)
    ws = _sheet(wb, sheet or None)
    for part in cell_range.split(","):
        min_c, min_r, max_c, max_r = range_boundaries(part.strip())
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                font_kw = {}
                f = cell.font
                if bold is not None:
                    font_kw["bold"] = bold
                if italic is not None:
                    font_kw["italic"] = italic
                if font_size is not None:
                    font_kw["size"] = font_size
                if font_color:
                    font_kw["color"] = font_color.lstrip("#").upper()
                if font_kw:
                    cell.font = Font(
                        bold=font_kw.get("bold", f.bold),
                        italic=font_kw.get("italic", f.italic),
                        size=font_kw.get("size", f.size),
                        color=font_kw.get("color", f.color),
                        name=f.name,
                    )
                if bg_color:
                    hexv = bg_color.lstrip("#").upper()
                    cell.fill = PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")
                if align:
                    cell.alignment = Alignment(horizontal=align, vertical=cell.alignment.vertical)
                if number_format:
                    cell.number_format = number_format
        if col_width is not None:
            for c in range(min_c, max_c + 1):
                ws.column_dimensions[get_column_letter(c)].width = col_width
    atomic_save(wb.save, path)
    return f"{cell_range} に書式を設定しました"


@tool
def excel_format_text(
    filename: str,
    keywords: list[str],
    sheet: str = "",
    color: str = "",
    bold: Optional[bool] = None,
    italic: Optional[bool] = None,
    underline: Optional[bool] = None,
    font_size: Optional[float] = None,
    font: str = "",
    highlight: str = "",
) -> str:
    """Excelのシートから特定の語句(キーワード)を含むセルを探し、文字に色・太字などの書式を付ける。
    「重要な用語を赤字にして」「◯◯のセルに蛍光ペンを」のような依頼に使う。Excelの文字書式はセル単位のため、
    セル内の一部だけでなくそのセルの文字全体に適用される。sheet省略時はアクティブシートが対象。
    colorは"#RRGGBB"形式(赤字なら"#FF0000")。bold/italic/underlineはtrueで付け、falseで外す。
    font_sizeは文字サイズ(pt)、fontはフォント名(例: "游ゴシック")。
    highlightは蛍光ペンの代わりにセルの塗りつぶし色になる("#RRGGBB"形式。黄色なら"#FFFF00")。
    数式("="で始まるセル)は対象外。セル範囲が分かっているときは excel_format を使う。"""
    keywords, error = validate_format_args(keywords, color, bold, italic, underline, font_size, font, highlight)
    if error:
        return error
    wb, path = _open(filename)
    if sheet and sheet not in wb.sheetnames:
        return f"エラー: シート「{sheet}」がありません。あるシート: {', '.join(wb.sheetnames)}"
    ws = wb[sheet] if sheet else wb.active
    hexv = color.lstrip("#").upper() if color else ""
    hl_hex = highlight.lstrip("#").upper() if highlight else ""
    refs: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or v.startswith("="):
                continue
            if not any(kw in v for kw in keywords):
                continue
            f = copy(cell.font)  # 既存のフォント設定を保ったまま指定分だけ変える
            if hexv:
                f.color = Font(color=hexv).color
            if bold is not None:
                f.bold = bold
            if italic is not None:
                f.italic = italic
            if underline is not None:
                f.underline = "single" if underline else None
            if font_size is not None:
                f.size = float(font_size)
            if font:
                f.name = font
            cell.font = f
            if hl_hex:
                cell.fill = PatternFill(start_color=hl_hex, end_color=hl_hex, fill_type="solid")
            refs.append(cell.coordinate)
    if not refs:
        return (f"「{'」「'.join(keywords)}」を含むセルはシート「{ws.title}」にありませんでした。"
                "excel_readで実際の表記を確認してください")
    atomic_save(wb.save, path)
    shown = ", ".join(refs[:10]) + (" ほか" if len(refs) > 10 else "")
    return (f"シート「{ws.title}」の{len(refs)}セル({shown})に"
            f"書式({describe_format(color, bold, italic, underline, font_size, font, highlight)})を適用しました")


def _rows_to_ranges(rows: list[int], max_col: int) -> str:
    """連続する行番号をまとめて、excel_formatに渡せる範囲文字列を作る。"""
    last_col = get_column_letter(max_col)
    parts = []
    start = prev = rows[0]
    for r in rows[1:] + [None]:
        if r is not None and r == prev + 1:
            prev = r
            continue
        parts.append(f"A{start}:{last_col}{start}" if start == prev else f"A{start}:{last_col}{prev}")
        if r is not None:
            start = prev = r
    return ",".join(parts)


@tool
def excel_query(filename: str, column: str, op: str, value: Any, sheet: str = "", header_row: int = 1) -> str:
    """Excelの表から条件に合う行を探し、行番号を返す。全データを読まずに対象行を特定できる。
    「売上が100万円を超える行に色を付けて」のような依頼では、まずこれで行番号を特定し、
    結果に含まれる範囲文字列をそのままexcel_formatに渡すとよい。
    column: 見出し行の列名(例: "売上")または列記号(例: "B")。header_rowは見出しの行番号(既定1)。
    op: "=" / "!=" / ">" / ">=" / "<" / "<=" / "contains"(部分一致)。
    注意: 数式セルは計算結果が保存されている場合のみ判定できる(判定できないセルは対象外になる)。"""
    path = resolve_workspace_path(filename, must_exist=True)
    if path.suffix != ".xlsx":
        return "エラー: Excelブックは .xlsx ファイルを指定してください"
    wb = load_workbook(str(path), data_only=True)  # 数式はキャッシュ済みの計算結果で比較する
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    if op not in ("=", "==", "!=", ">", ">=", "<", "<=", "contains"):
        return 'エラー: opは "=" "!=" ">" ">=" "<" "<=" "contains" のいずれかを指定してください'
    # 列名→列番号の解決(見出し行を優先し、見つからなければ列記号として解釈)
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            headers[str(v).strip()] = c
    if column.strip() in headers:
        col_idx = headers[column.strip()]
    elif column.strip().isascii() and column.strip().isalpha() and len(column.strip()) <= 3:
        col_idx = range_boundaries(f"{column.strip().upper()}1")[0]
    else:
        return f"エラー: 列「{column}」が見出し行{header_row}に見つかりません。ある列名: {', '.join(list(headers)[:20])}"

    def _match(v: Any) -> bool:
        if op == "contains":
            return v is not None and str(value) in str(v)
        try:
            a, b = float(v), float(value)  # 数値として比較できるなら数値で
        except (TypeError, ValueError):
            if op in ("=", "=="):
                return v is not None and str(v).strip() == str(value).strip()
            if op == "!=":
                return v is None or str(v).strip() != str(value).strip()
            return False  # 大小比較は数値のみ
        return {"=": a == b, "==": a == b, "!=": a != b, ">": a > b, ">=": a >= b, "<": a < b, "<=": a <= b}[op]

    hits = [r for r in range(header_row + 1, ws.max_row + 1) if _match(ws.cell(row=r, column=col_idx).value)]
    if not hits:
        return f"「{column}」が {op} {value} に該当する行はありませんでした"
    shown = ", ".join(str(r) for r in hits[:100]) + ("..." if len(hits) > 100 else "")
    lines = [f"「{column}」が {op} {value} の行: {shown} ({len(hits)}件)"]
    if len(hits) <= 50:
        lines.append(f'excel_formatにそのまま使える範囲: "{_rows_to_ranges(hits, ws.max_column)}"')
    return "\n".join(lines)


@tool
def excel_add_sheet(filename: str, sheet_name: str) -> str:
    """Excelブックに新しいシートを追加する。"""
    wb, path = _open(filename)
    if sheet_name in wb.sheetnames:
        return f"シート「{sheet_name}」は既に存在します"
    wb.create_sheet(sheet_name)
    atomic_save(wb.save, path)
    return f"シート「{sheet_name}」を追加しました"


@tool
def excel_add_chart(
    filename: str,
    chart_type: str,
    data_range: str,
    categories_range: str = "",
    anchor: str = "",
    sheet: str = "",
    title: str = "",
    x_title: str = "",
    y_title: str = "",
    titles_from_data: bool = True,
    stacked: bool = False,
    error_bars_range: str = "",
) -> str:
    """Excel上で編集できるグラフ(ネイティブのグラフ)をシートに追加する。画像ではないので、
    あとからExcelで種類や色を変更でき、参照元のセルの値を直すとグラフも自動で追従する。
    Excelのグラフはこのツールで作ること(run_pythonで画像を貼るのはWordに載せる図のときだけ)。
    chart_type: bar(縦棒) / bar_horizontal(横棒) / line(折れ線) / pie(円)。
    data_rangeは数値の範囲、categories_rangeは項目名の範囲をA1形式で指定する(例: "B1:C5", "A2:A5")。
    titles_from_data=Trueなら data_range の1行目を系列名として扱う(その場合は見出し行を含めること)。
    例: 表がA1:C5(A列=月, B列=売上, C列=利益, 1行目が見出し)のとき、
    data_range="B1:C5", categories_range="A2:A5", anchor="E2"。
    anchorはグラフの左上を置くセル(省略時は表の右隣)。stacked=Trueで積み上げ棒。
    pieは最初の1系列だけが使われる。
    error_bars_range: エラーバー(誤差範囲)にする値の範囲。標準偏差や標準誤差の列を指定すると、
    平均の棒に±のひげが付く(例: 平均がB列・標準偏差がC列なら data_range="B1:B5",
    error_bars_range="C2:C5")。1系列のグラフにだけ付けられる。"""
    type_map = {"bar": BarChart, "bar_horizontal": BarChart, "line": LineChart, "pie": PieChart}
    if chart_type not in type_map:
        return f"エラー: chart_typeは {' / '.join(type_map)} のいずれかを指定してください"
    wb, path = _open(filename)
    ws = _sheet(wb, sheet or None)
    try:
        bounds = range_boundaries(data_range)
    except ValueError:
        return f"エラー: data_range「{data_range}」はA1形式(例: B1:C5)で指定してください"
    min_col, min_row, max_col, max_row = bounds
    if max_row - min_row < 1 and max_col - min_col < 1:
        return "エラー: data_rangeには複数のセルを含む範囲を指定してください(例: B1:B5)"

    chart = type_map[chart_type]()
    if chart_type == "bar_horizontal":
        chart.type = "bar"
    if stacked and chart_type in ("bar", "bar_horizontal"):
        chart.grouping = "stacked"
        chart.overlap = 100  # これが無いとExcelで積み上がって見えない
    if title:
        chart.title = title
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title
    chart.add_data(Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row),
                   titles_from_data=titles_from_data)
    if categories_range:
        try:
            c_min_col, c_min_row, c_max_col, c_max_row = range_boundaries(categories_range)
        except ValueError:
            return f"エラー: categories_range「{categories_range}」はA1形式(例: A2:A5)で指定してください"
        chart.set_categories(Reference(ws, min_col=c_min_col, min_row=c_min_row,
                                       max_col=c_max_col, max_row=c_max_row))
    # 既定色はUIと同じ配色に揃える(色を指定しないとExcelのテーマ色になり、プレビューと見た目がずれる)。
    # 円グラフは「系列」ではなく「項目」ごとに色を変えるので、データ点単位で色を付ける
    if chart_type == "pie":
        n_points = max_row - min_row + (0 if titles_from_data else 1)
        chart.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=CHART_COLORS[i % len(CHART_COLORS)]))
            for i in range(n_points)
        ]
    else:
        for i, s in enumerate(chart.series):
            s.graphicalProperties.solidFill = CHART_COLORS[i % len(CHART_COLORS)]
    note = ""
    if error_bars_range:
        if chart_type == "pie":
            return "エラー: 円グラフにはエラーバーを付けられません"
        try:
            range_boundaries(error_bars_range)
        except ValueError:
            return f"エラー: error_bars_range「{error_bars_range}」はA1形式(例: C2:C5)で指定してください"
        if len(chart.series) > 1:
            note = "\n注意: エラーバーは最初の系列にだけ付けました(複数系列には対応していません)"
        ref = f"'{ws.title}'!{_absolute(error_bars_range)}"
        # errValType='cust' = 値をセルから取る。plus/minusに同じ範囲を渡して±のひげにする
        chart.series[0].errBars = ErrorBars(
            errBarType="both", errValType="cust",
            plus=NumDataSource(NumRef(f=ref)), minus=NumDataSource(NumRef(f=ref)),
        )
    if not anchor:
        anchor = f"{get_column_letter(max_col + 2)}{min_row}"  # 表の右隣に置く
    ws.add_chart(chart, anchor)
    atomic_save(wb.save, path)
    where = f"シート「{ws.title}」の{anchor}"
    bars = "エラーバー付きの" if error_bars_range else ""
    return (f"{where}に{bars}グラフ({chart_type})を追加しました。Excel上で編集できるグラフなので、"
            f"参照元のセル({data_range})を変更すればグラフも変わります{note}")


EXCEL_TOOLS = [excel_create, excel_read, excel_query, excel_write_cells, excel_write_rows, excel_format,
               excel_format_text, excel_add_sheet, excel_add_chart]
